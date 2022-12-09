#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@author:TXU
@file:report_xlsx
@time:2022/12/6
@email:tao.xu2008@outlook.com
@description: 生成Excel报告
"""
import os
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from fio_report.models import ExcelReportSettings


class ReportXlsx(object):
    """生成EXCEL报告"""
    def __init__(self, output_path, json_data):
        self.output_path = output_path
        self.json_data = json_data
        self.settings = ExcelReportSettings()
        self.row_count = 0

        self.wb = Workbook(write_only=False)
        self.data_ws = self.wb.create_sheet(self.settings.data_sheet_title, self.settings.data_sheet_index)
        self.data_ws.append(self.settings.data_column_title)
        self.chart_ws = self.wb.create_sheet(self.settings.chart_sheet_title, self.settings.chart_sheet_index)

    @staticmethod
    def get_item_index(items, key):
        """
        获取 key在items中的位置
        :param items:
        :param key:
        :return:
        """
        for idx, item in enumerate(items):
            if item == key:
                return idx+1
        raise Exception("{}中没找到:{}".format(items, key))

    def reset_col(self):
        for idx, col in enumerate(self.data_ws.columns):
            letter = get_column_letter(idx + 1)  # 列字母
            collen = max([len(str(c.value).encode()) for c in col])  # 获取这一列长度的最大值
            self.data_ws.column_dimensions[letter].width = collen * 1.2 + 4  # 也就是列宽为最大长度*1.2 可以自己调整

    def write_data_sheet(self):
        for bs_data in self.json_data:
            for data in bs_data['data']:
                row = [
                    # 测试 重要参数
                    data['jobname'],
                    data['rw'],
                    data['iodepth'],
                    data['numjobs'],
                    data['bs'],
                    # 写 结果
                    0,
                    0,
                    0,
                    # 读 结果
                    0,
                    0,
                    0,
                ]
                for result in data["result"]:
                    idx_start = 5 if result["type"] == "write" else 8
                    row[idx_start+0] = result['bw_mb']
                    row[idx_start+1] = result['iops']
                    row[idx_start+2] = result['lat_ms']
                logger.debug(row)
                self.data_ws.append(row)
                self.row_count += 1
        self.data_ws.auto_filter.ref = f"A1:K{self.row_count+1}"
        self.data_ws.auto_filter.add_filter_column(0, [])
        self.data_ws.auto_filter.add_sort_condition(f"F2:F{self.row_count+1}")
        self.data_ws.auto_filter.add_sort_condition(f"G2:G{self.row_count+1}")
        self.data_ws.auto_filter.add_sort_condition(f"H2:H{self.row_count+1}")

    def bar_chart(self, key, x_title=None, y_title=None):
        """
        创建bar chart
        :param key:
        :param x_title:
        :param y_title:
        :return:
        """
        chart = BarChart()
        chart.type = "bar"
        chart.style = 13
        # chart.overlap = 100
        chart.height = 6+(self.row_count*1.2)
        chart.legend.position = "b"  # 下方 显示图例
        chart.title = "FIO性能对比 - {}".format(key.split("-")[0].upper())
        chart.y_axis.title = y_title or 'Test number'
        chart.x_axis.title = x_title or 'Test Job Name'
        col_index = self.get_item_index(self.settings.data_column_title, key)
        data1 = Reference(self.data_ws, min_row=1, max_row=self.row_count+1, min_col=col_index, max_col=col_index)
        data2 = Reference(self.data_ws, min_row=1, max_row=self.row_count+1, min_col=col_index+3, max_col=col_index+3)
        cats = Reference(self.data_ws, min_row=2, max_row=self.row_count+1, min_col=1)
        chart.add_data(data1, titles_from_data=True)
        chart.add_data(data2, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        # 显示数据标签
        for s in chart.series:
            s.dLbls = DataLabelList()
            s.dLbls.showVal = True

        return chart

    def bar_chart_bw(self):
        return self.bar_chart("bw-write", y_title="Test number(MiB)")

    def bar_chart_iops(self):
        return self.bar_chart("iops-write")

    def bar_chart_lat(self):
        return self.bar_chart("latency-write", y_title="Test number(ms)")

    def create_xlsx_file(self):
        # 创建数据统计表
        self.write_data_sheet()
        self.reset_col()

        # 创建性能对比图
        self.chart_ws.add_chart(self.bar_chart_bw(), "A1")
        self.chart_ws.add_chart(self.bar_chart_iops(), "I1")
        self.chart_ws.add_chart(self.bar_chart_lat(), "Q1")
        self.wb.save(os.path.join(os.path.abspath(self.output_path), "report.xlsx"))

        # from copy import deepcopy
        #
        # chart2 = deepcopy(chart1)
        # chart2.style = 11
        # chart2.type = "bar"
        # chart2.title = "Horizontal Bar Chart"
        #
        # self.chart_ws.add_chart(chart2, "G10")
        #
        # chart3 = deepcopy(chart1)
        # chart3.type = "col"
        # chart3.style = 12
        # chart3.grouping = "stacked"
        # chart3.overlap = 100
        # chart3.title = 'Stacked Chart'
        #
        # self.chart_ws.add_chart(chart3, "A27")
        #
        # chart4 = deepcopy(chart1)
        # chart4.type = "bar"
        # chart4.style = 13
        # chart4.grouping = "percentStacked"
        # chart4.overlap = 100
        # chart4.title = 'Percent Stacked Chart'
        #
        # self.chart_ws.add_chart(chart4, "G27")


if __name__ == '__main__':
    pass
